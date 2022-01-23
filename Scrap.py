from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup as BS
from time import sleep
import requests
import os
import re
import getpass
import openpyxl
import time


start_time = time.time()
print("Getting ready!")

options = webdriver.ChromeOptions()
options.headless = True

username = getpass.getuser()
path = 'C:\\Users\\' + username + '\\Desktop'
os.chdir(path)
if not os.path.isdir("NGO_Webscraping"):
    os.mkdir("NGO_Webscraping")
os.chdir('NGO_Webscraping')

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = 'NGOs in J&K'

sheet.cell(1, 1).value = "Sr No."
sheet.cell(1, 2).value = "Name of NGO/VO"
sheet.cell(1, 3).value = "Email"
sheet.cell(1, 4).value = "Ph No."
sheet.cell(1, 5).value = "City"
sheet.cell(1, 6).value = "Address"
sheet.cell(1, 7).value = "Sectors"
sheet.cell(1, 8).value = "Website"
sheet.cell(1, 9).value = "Opeartional_States"
sheet.cell(1, 10).value = "Opeartional_District"
sheet.cell(1, 11).value = "Date of establishment"
sheet.cell(1, 12).value = "ID"

browser = webdriver.Chrome(options=options)

link_a = 'https://ngodarpan.gov.in/index.php/home/statewise_ngo/1231/1/'
link_b = '?per_page=100'
current_page = 1

url = link_a + str(current_page) + link_b

# browser.get(url)
src = requests.get(url)
soup = BS(src.text, 'html.parser')

total_page = int(len(soup.find('ul', {'class': 'pagination'}).findAll('li')))

cross_selector = '#ngo_info_modal > div.modal-dialog.modal-lg > div > div.modal-header > button > span'

ngo_count = 0

for current_page in range(1, total_page):

    url = link_a + str(current_page) + link_b
    src = requests.get(url)
    soup = BS(src.text, 'html.parser')

    browser.get(url)

    table = soup.find('table', {'class': 'table table-striped table-bordered table-hover Tax'}).find('tbody')
    iterator = table.findChildren('tr')

    for ngo in iterator:

        sleep(0.5)

        data = ngo.findAll('td')
        link = data[1]
        linkText = link.text

        srNo = int(data[0].text) + 100 * (current_page - 1)
        name = data[1].text
        sector = data[3].text

        print("Extracting NGO number:", srNo)

        selector = 'body > div:nth-child(21) > div.container > div.row > div > div > div.ibox-content > table > tbody > tr:nth-child(' + str(
            data[0].text) + ') > td:nth-child(2) > a'
        # print(linkText.strip())

        try:
            pdf = browser.find_element_by_css_selector(selector)

            pdf.click()

            try:

                pleasewait = WebDriverWait(browser, 60).until(
                    EC.invisibility_of_element((By.CSS_SELECTOR, 'body > div.blockUI.blockMsg.blockPage > h4')))

                id = browser.find_element_by_id('UniqueID')
                id = id.text
                # print(id)

                email = browser.find_element_by_id('email_n')
                email = email.text
                email = email.replace('(at)', '@').replace('[dot]', '.').lower()
                # print(email)

                city = browser.find_element_by_id('city')
                city = city.text
                # print(city)

                addr = browser.find_element_by_id('address')
                addr = addr.text
                # print(addr)

                phno = browser.find_element_by_id('mobile_n')
                phno = phno.text
                # print(phno)

                state = browser.find_element_by_id('operational_states')
                state = state.text
                # print(state)

                area = browser.find_element_by_id('operational_district')
                area = area.text
                # print(area)

                website = browser.find_element_by_id('ngo_web_url')
                website = website.text
                # print(website)

                date = browser.find_element_by_id('ngo_reg_date')
                date = date.text
                # print(date)

                cross = browser.find_element_by_css_selector(cross_selector)
                cross.click()

                sheet.cell(srNo + 1, 1).value = srNo
                sheet.cell(srNo + 1, 2).value = name
                sheet.cell(srNo + 1, 3).value = email
                sheet.cell(srNo + 1, 4).value = phno
                sheet.cell(srNo + 1, 5).value = city
                sheet.cell(srNo + 1, 6).value = addr
                sheet.cell(srNo + 1, 7).value = sector
                sheet.cell(srNo + 1, 8).value = website
                sheet.cell(srNo + 1, 9).value = state
                sheet.cell(srNo + 1, 10).value = area
                sheet.cell(srNo + 1, 11).value = date
                sheet.cell(srNo + 1, 12).value = id

                wb.save('NGOs_J&K.xlsx')

                ngo_count += 1

            except Exception as e:
                print(e)
                print("Unable to extract details for NGO number:", data[0].text, "on the page number:",
                      str(current_page))
                sheet.cell(srNo + 1, 1).value = "CODE FAILURE"
                continue

            # sleep(5)

            '''try:
                element = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'email_n')))
                while len(element.text)==0:
                    element = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'email_n')))
                email =  element.text

                print(email)

                cross = browser.find_element_by_css_selector(cross_selector)
                cross.click()


            except Exception as e:
                print(e)
                print("Timed out waiting for page to load.")'''

            '''try:
                element_email = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'email_n')))
                while len(element_email.text) == 0:
                    element_email = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'email_n')))
                email = element_email.text
                email = email.replace('(at)', '@').replace('[dot]', '.')
                print(email)

                try:
                    element_city = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'city')))
                    while len(element_city.text)==0:
                        element_city = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'city')))
                    city = element_city.text
                    print(city)

                    try:
                        element_phno = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'mobile_n')))
                        while len(element_phno.text) == 0:
                            element_phno = WebDriverWait(browser, 3).until(EC.presence_of_element_located((By.ID, 'mobile_n')))
                        phno = element_phno.text
                        print(phno)

                        try:
                            element_addr = WebDriverWait(browser, 3).until(
                                EC.presence_of_element_located((By.ID, 'address')))
                            while len(element_addr.text) == 0:
                                element_addr = WebDriverWait(browser, 3).until(
                                    EC.presence_of_element_located((By.ID, 'address')))
                            addr = element_addr.text
                            print(addr)

                            try:
                                element_website = WebDriverWait(browser, 3).until(
                                    EC.presence_of_element_located((By.ID, 'ngo_web_url')))
                                while len(element_addr.text) == 0:
                                    element_website = WebDriverWait(browser, 3).until(
                                        EC.presence_of_element_located((By.ID, 'ngo_web_url')))
                                website = element_website.text
                                print(website)


                                try:
                                    pleasewait = WebDriverWait(browser, 60).until(
                                    EC.invisibility_of_element((By.CSS_SELECTOR, 'body > div.blockUI.blockMsg.blockPage > h4')))
                                    cross = browser.find_element_by_css_selector(cross_selector)
                                    cross.click()
                                    print('------------------------------------')

                                except:
                                    print("Time out in getting cross button :(")

                            except:
                                print("Time out in getting NGOs website url :(")

                        except:
                            print("Time out in getting address :(")

                    except:
                        print("Time out in getting phone number :(")

                except:
                    print("Time out in city :(")

                try:
                    cross_element = WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cross_selector)))
                    cross = browser.find_element_by_css_selector(cross_selector)
                    cross.click()
                except:
                    print("Cross taking long :(")


            except Exception as e:
                print(e)
                print("Timed out waiting for page to load.")'''



        except:
            print("CSS selector error in NGO number:", data[0].text, "on page number:", str(current_page))
            sheet.cell(srNo + 1, 1).value = "CSS CODE FAILURE"
            continue

    # break

wb.save('NGOs_J&K.xlsx')
wb.close()

time_taken = time.time() - start_time
print("Extracted", ngo_count, "NGO details in:", time_taken / 60, "minutes.")
browser.close()
